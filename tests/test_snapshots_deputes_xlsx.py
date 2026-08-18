import pytest

from votes_parlementaires.an.build import scrutins_csv
from votes_parlementaires.config import AN_LEGISLATURE_DEFAULT
from votes_parlementaires.snapshots.deputes_xlsx import xlsx_sheet_name

pytestmark_integration = pytest.mark.skipif(
    not scrutins_csv(AN_LEGISLATURE_DEFAULT).exists(),
    reason="Nécessite les CSV construits (python -m votes_parlementaires.an.build).",
)


def test_xlsx_sheet_name_truncates_to_31_chars():
    long_name = "Un Nom De Député Vraiment Beaucoup Trop Long Pour Excel"
    name = xlsx_sheet_name(long_name, set())
    assert len(name) <= 31


def test_xlsx_sheet_name_strips_forbidden_characters():
    name = xlsx_sheet_name("Jean/Paul: Dupont*Martin?", set())
    assert not any(c in name for c in ":\\/?*[]")


def test_xlsx_sheet_name_dedupes_collisions():
    used = set()
    a = xlsx_sheet_name("Jean Dupont", used)
    b = xlsx_sheet_name("Jean Dupont", used)
    assert a != b
    assert len(b) <= 31


@pytestmark_integration
def test_generate_produces_one_sheet_per_depute_plus_reference_sheets(tmp_path):
    import openpyxl

    from votes_parlementaires.snapshots.deputes_xlsx import generate

    out = generate(["17", "79"], legislature=AN_LEGISLATURE_DEFAULT, out=tmp_path / "deputes-17-79.xlsx")
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == "Sommaire"
    assert "Catégories" in wb.sheetnames
    assert len(wb.sheetnames) == 2 + 8  # Sommaire + Catégories + 8 député·e·s

    for name in wb.sheetnames:
        assert list(wb[name].tables.keys()), f"aucune table Excel sur l'onglet {name!r}"

"""Génère un classeur Excel (données figées) listant les député·e·s de
départements donnés, un onglet par député·e, avec le détail de leurs
scrutins (date, texte de loi, catégorie, position, résultat). Miroir en
tableur du contenu des pages HTML figées (`snapshots/deputes.py`) : mêmes
données, mêmes catégories, même regroupement par texte de loi — mais sous
forme de tableau filtrable nativement dans Excel plutôt que d'accordéons.

Usage :
    python -m votes_parlementaires.snapshots.deputes_xlsx --departements 17 79
"""

import argparse
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from votes_parlementaires.an.meta import read_build_date
from votes_parlementaires.an.taxonomy import CATEGORIES, CATEGORY_LABELS
from votes_parlementaires.config import an_snapshots_dir
from votes_parlementaires.snapshots.deputes import (
    MOIS_FULL,
    fmt_date,
    load_deputes_data,
    to_roman,
)

POSITION_LABELS = {"pour": "Pour", "contre": "Contre", "abstention": "Abstention", "nonVotant": "Non-votant"}

TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(color="666666", italic=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E4C7A")
BOLD = Font(bold=True)

VOTE_COLUMNS = ["Date", "Texte de loi", "Catégorie", "Détail du scrutin", "Position", "Résultat", "Par délégation"]


def xlsx_sheet_name(nom_complet: str, used: set[str]) -> str:
    """Nom d'onglet Excel valide (<=31 caractères, sans : \\ / ? * [ ]),
    dédupliqué si collision (peu probable mais possible avec des homonymes)."""
    name = re.sub(r"[:\\/?*\[\]]", "", nom_complet).strip()[:31]
    base, n = name, 2
    while name in used:
        suffix = f" ({n})"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def autosize_columns(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def write_sommaire(wb: Workbook, dd, snapshot_label: str) -> dict[str, str]:
    ws = wb.active
    ws.title = "Sommaire"
    ws["A1"] = "Député·e·s"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = snapshot_label
    ws["A2"].font = SUBTITLE_FONT

    headers = ["Nom", "Groupe", "Département", "N° circ.", "Participation", "Pour", "Contre", "Abstention", "Non-votant", "Onglet"]
    header_row = 4
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    used_sheet_names: set[str] = set()
    sheet_by_ref: dict[str, str] = {}
    row = header_row + 1
    for dep_row in dd.deps_df.itertuples():
        entry = dd.entries[dep_row.acteur_ref]
        sheet_name = xlsx_sheet_name(dep_row.nom_complet, used_sheet_names)
        sheet_by_ref[dep_row.acteur_ref] = sheet_name
        pct = round(entry["recorded"] / entry["possible"] * 100) if entry["possible"] else None

        ws.cell(row=row, column=1, value=f"{dep_row.civ} {dep_row.nom_complet}")
        ws.cell(row=row, column=2, value=dep_row.groupe_libelle or "Non inscrit")
        ws.cell(row=row, column=3, value=dep_row.departement)
        ws.cell(row=row, column=4, value=int(dep_row.num_circonscription))
        ws.cell(row=row, column=5, value=(pct / 100 if pct is not None else None))
        if pct is not None:
            ws.cell(row=row, column=5).number_format = "0%"
        ws.cell(row=row, column=6, value=entry["stats"]["pour"])
        ws.cell(row=row, column=7, value=entry["stats"]["contre"])
        ws.cell(row=row, column=8, value=entry["stats"]["abstention"])
        ws.cell(row=row, column=9, value=entry["stats"]["nonVotant"])
        link_cell = ws.cell(row=row, column=10, value=sheet_name)
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="2E4C7A", underline="single")
        row += 1

    table = Table(displayName="Sommaire", ref=f"A{header_row}:J{row - 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    autosize_columns(ws, [26, 30, 20, 9, 13, 7, 7, 11, 11, 20])
    return sheet_by_ref


def write_categories_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Catégories")
    ws["A1"] = "Taxonomie des scrutins"
    ws["A1"].font = TITLE_FONT
    headers = ["Catégorie", "Identifiant", "Description", "Couleur"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, c in enumerate(CATEGORIES, start=4):
        ws.cell(row=i, column=1, value=c["label"])
        ws.cell(row=i, column=2, value=c["id"])
        ws.cell(row=i, column=3, value=c["description"])
        color_cell = ws.cell(row=i, column=4, value=c["color"])
        color_cell.fill = PatternFill("solid", fgColor=c["color"].lstrip("#").upper())
    table = Table(displayName="Categories", ref=f"A3:D{3 + len(CATEGORIES)}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    autosize_columns(ws, [34, 30, 60, 10])


def write_depute_sheet(wb: Workbook, sheet_name: str, table_index: int, dep_row, entry: dict, groups) -> None:
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = f"{dep_row.civ} {dep_row.nom_complet}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{dep_row.groupe_libelle or 'Non inscrit'} — {dep_row.departement} (circonscription n°{dep_row.num_circonscription})"
    ws["A2"].font = SUBTITLE_FONT

    pct = round(entry["recorded"] / entry["possible"] * 100) if entry["possible"] else None
    if pct is not None:
        ws["A3"] = f"{pct} % de participation ({entry['recorded']} / {entry['possible']} scrutins depuis le {entry['since']})"
    else:
        ws["A3"] = "Participation non calculable (aucun vote enregistré)"

    header_row = 5
    for col, label in enumerate(VOTE_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row = header_row + 1
    for v in entry["votes"]:
        texte, cat_id = groups.groups[v["gi"]]
        ws.cell(row=row, column=1, value=date.fromisoformat(v["iso"]) if v["iso"] else None)
        ws.cell(row=row, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=2, value=texte)
        ws.cell(row=row, column=3, value=CATEGORY_LABELS.get(cat_id, cat_id))
        ws.cell(row=row, column=4, value=v["a"] or v["t"])
        ws.cell(row=row, column=5, value=POSITION_LABELS.get(v["p"], v["p"]))
        ws.cell(row=row, column=6, value=v["r"])
        ws.cell(row=row, column=7, value="Oui" if v["g"] else "Non")
        row += 1

    if row > header_row + 1:
        table = Table(displayName=f"DeputeVotes{table_index}", ref=f"A{header_row}:G{row - 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    autosize_columns(ws, [12, 55, 30, 55, 12, 10, 12])
    for r in (ws["A2"], ws["A3"]):
        r.alignment = Alignment(wrap_text=False)


def generate(departements: list[str], legislature: int | None = None, out: Path | None = None) -> Path:
    dd = load_deputes_data(departements, legislature)
    codes = [str(c) for c in departements]

    build_date = read_build_date(dd.legislature)
    if build_date is None:
        snapshot_label = "Données figées (date de construction inconnue — reconstruire via build.py) — source data.assemblee-nationale.fr"
    else:
        snapshot_label = (
            f"Données figées au {fmt_date(build_date.isoformat(), MOIS_FULL)} — "
            f"Assemblée nationale, {to_roman(dd.legislature)}e législature — source data.assemblee-nationale.fr"
        )

    wb = Workbook()
    sheet_by_ref = write_sommaire(wb, dd, snapshot_label)
    write_categories_sheet(wb)
    for i, dep_row in enumerate(dd.deps_df.itertuples()):
        write_depute_sheet(wb, sheet_by_ref[dep_row.acteur_ref], i, dep_row, dd.entries[dep_row.acteur_ref], dd.groups)

    if out is None:
        out = an_snapshots_dir(dd.legislature) / f"deputes-{'-'.join(codes)}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Génère un classeur Excel (un onglet par député·e) listant les député·e·s de départements donnés.")
    parser.add_argument("--departements", nargs="+", default=["17", "79"], help="Numéros de département (ex: 17 79).")
    parser.add_argument("--legislature", type=int, default=None, help="Forcer une législature (par défaut: la plus récente déjà construite).")
    parser.add_argument("--out", type=Path, default=None, help="Chemin de sortie (par défaut: data/snapshots/an/<législature>/deputes-<deps>.xlsx).")
    args = parser.parse_args()

    out = generate(args.departements, legislature=args.legislature, out=args.out)
    print(f"Classeur généré : {out}")


if __name__ == "__main__":
    main()

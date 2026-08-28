"""Génère un classeur Excel (données figées) listant les sénateur·rice·s de
départements donnés, un onglet par sénateur·rice, avec le détail de leurs
scrutins depuis le début de leur mandat. Miroir en tableur de la page HTML
figée (`snapshots/senateurs.py`) — même structure que `snapshots/deputes_xlsx.py`
côté Assemblée nationale.

Usage :
    python -m votes_parlementaires.snapshots.senateurs_xlsx --departements 17 79
"""

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from votes_parlementaires.an.taxonomy import CATEGORY_LABELS
from votes_parlementaires.config import senat_processed_dir
from votes_parlementaires.snapshots.deputes import GroupRegistry
from votes_parlementaires.snapshots.deputes_xlsx import (
    HEADER_FILL,
    HEADER_FONT,
    POSITION_LABELS,
    SUBTITLE_FONT,
    TITLE_FONT,
    autosize_columns,
    write_categories_sheet,
    xlsx_sheet_name,
)
from votes_parlementaires.senat.categorize import load_categorie_map
from votes_parlementaires.an.categorize import distinct_textes
from votes_parlementaires.senat.data import SenatData
from votes_parlementaires.snapshots.senateurs import build_senateur_entry

VOTE_COLUMNS = ["Date", "Texte de loi", "Catégorie", "Détail du scrutin", "Position", "Résultat"]


def write_sommaire(wb: Workbook, sen_df, entries: dict, snapshot_label: str) -> dict[str, str]:
    ws = wb.active
    ws.title = "Sommaire"
    ws["A1"] = "Sénateur·rice·s"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = snapshot_label
    ws["A2"].font = SUBTITLE_FONT

    headers = ["Nom", "Groupe", "Département", "Depuis", "Participation", "Pour", "Contre", "Abstention", "Non-votant", "Onglet"]
    header_row = 4
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    used_sheet_names: set[str] = set()
    sheet_by_ref: dict[str, str] = {}
    row = header_row + 1
    for sen_row in sen_df.itertuples():
        entry = entries[sen_row.senateur_ref]
        sheet_name = xlsx_sheet_name(sen_row.nom_complet, used_sheet_names)
        sheet_by_ref[sen_row.senateur_ref] = sheet_name
        pct = round(entry["recorded"] / entry["possible"] * 100) if entry["possible"] else None

        ws.cell(row=row, column=1, value=f"{sen_row.civ} {sen_row.nom_complet}")
        ws.cell(row=row, column=2, value=sen_row.groupe_libelle or "Non inscrit")
        ws.cell(row=row, column=3, value=sen_row.departement)
        ws.cell(row=row, column=4, value=entry["since"] or "")
        ws.cell(row=row, column=5, value=(pct / 100 if pct is not None else None))
        if pct is not None:
            ws.cell(row=row, column=5).number_format = "0%"
        ws.cell(row=row, column=6, value=entry["stats"]["pour"])
        ws.cell(row=row, column=7, value=entry["stats"]["contre"])
        ws.cell(row=row, column=8, value=entry["stats"]["abstention"])
        ws.cell(row=row, column=9, value=entry["stats"]["nonVotant"])
        link_cell = ws.cell(row=row, column=10, value=sheet_name)
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="2E4C7A", underline="single")
        row += 1

    table = Table(displayName="Sommaire", ref=f"A{header_row}:J{row - 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    autosize_columns(ws, [26, 22, 20, 14, 13, 7, 7, 11, 11, 20])
    return sheet_by_ref


def write_senateur_sheet(wb: Workbook, sheet_name: str, table_index: int, sen_row, entry: dict, groups: GroupRegistry) -> None:
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = f"{sen_row.civ} {sen_row.nom_complet}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{sen_row.groupe_libelle or 'Non inscrit'} — {sen_row.departement}"
    ws["A2"].font = SUBTITLE_FONT

    pct = round(entry["recorded"] / entry["possible"] * 100) if entry["possible"] else None
    if pct is not None:
        ws["A3"] = f"{pct} % de participation ({entry['recorded']} / {entry['possible']} scrutins depuis le {entry['since']})"
    else:
        ws["A3"] = "Participation non calculable (aucun vote enregistré)"

    header_row = 5
    for col, label in enumerate(VOTE_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row = header_row + 1
    for v in entry["votes"]:
        texte, cat_id = groups.groups[v["gi"]]
        ws.cell(row=row, column=1, value=date.fromisoformat(v["iso"]) if v["iso"] else None)
        ws.cell(row=row, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=2, value=texte)
        ws.cell(row=row, column=3, value=CATEGORY_LABELS.get(cat_id, cat_id))
        ws.cell(row=row, column=4, value=v["a"] or v["t"])
        ws.cell(row=row, column=5, value=POSITION_LABELS.get(v["p"], v["p"]))
        ws.cell(row=row, column=6, value=v["r"])
        row += 1

    if row > header_row + 1:
        table = Table(displayName=f"SenateurVotes{table_index}", ref=f"A{header_row}:F{row - 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    autosize_columns(ws, [12, 55, 30, 55, 12, 10])
    for r in (ws["A2"], ws["A3"]):
        r.alignment = Alignment(wrap_text=False)


def generate(departements: list[str], out: Path | None = None, groups: GroupRegistry | None = None) -> Path:
    from votes_parlementaires.snapshots.senateurs import load_senateurs_data

    if groups is None:
        data = SenatData()
        groups = GroupRegistry(distinct_textes(data.scrutins), load_categorie_map())

    sd = load_senateurs_data(departements, groups)
    codes = [str(c) for c in departements]

    snapshot_label = "Données figées — Sénat — source senat.fr"

    wb = Workbook()
    sheet_by_ref = write_sommaire(wb, sd.sen_df, sd.entries, snapshot_label)
    write_categories_sheet(wb)
    for i, sen_row in enumerate(sd.sen_df.itertuples()):
        write_senateur_sheet(wb, sheet_by_ref[sen_row.senateur_ref], i, sen_row, sd.entries[sen_row.senateur_ref], sd.groups)

    if out is None:
        out = senat_processed_dir() / f"senateurs-{'-'.join(codes)}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Génère un classeur Excel (un onglet par sénateur·rice) listant les sénateur·rice·s de départements donnés."
    )
    parser.add_argument("--departements", nargs="+", default=["17", "79"], help="Numéros de département (ex: 17 79).")
    parser.add_argument("--out", type=Path, default=None, help="Chemin de sortie (par défaut: data/processed/senat/senateurs-<deps>.xlsx).")
    args = parser.parse_args()

    out = generate(args.departements, out=args.out)
    print(f"Classeur généré : {out}")


if __name__ == "__main__":
    main()
